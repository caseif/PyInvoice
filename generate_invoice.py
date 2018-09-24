#!/usr/bin/python3

import argparse
from datetime import datetime, time
from functools import reduce
import itertools
from os import path

from openpyxl import load_workbook

DATE_FORMAT = '%m/%d/%Y'
TIME_FORMAT = '%H:%M'
DATETIME_FORMAT = "%s %s" % (DATE_FORMAT, TIME_FORMAT)
DATE_FORMAT_FILE_NAME = '%Y%m%d'
DATE_FORMAT_ISO = '%Y-%m-%d'

class Timesheet:
    def __init__(self, item, rate, ranges, opened, closed):
        self.item = item
        self.rate = rate
        self.ranges = ranges
        self.opened = opened
        self.closed = closed

    def get_owed(self):
        return reduce(lambda a, r: a + (r.duration().seconds / 3600 * self.rate), self.ranges, 0)

class TimesheetMetadata:
    def __init__(self, project, tracker, start_date, end_date):
        self.project = project
        self.tracker = tracker
        self.start_date = start_date
        self.end_date = end_date

class TimeRange:
    def __init__(self, start, end):
        self.start = start
        self.end = end

    def overlaps(self, range):
        return max(self.start, range.start) < min(self.end, range.end)

    def duration(self):
        return self.end - self.start

    def __str__(self):
        return "%s: %s - %s" % (self.start.strftime(DATE_FORMAT), self.start.time().strftime(TIME_FORMAT), self.end.time().strftime(TIME_FORMAT))

def parse_metadata(workbook):
    if not "META" in workbook:
        print('Timesheet document does not have a META sheet!')
        exit(1)

    meta_sheet = workbook["META"]

    project    = meta_sheet['B1'].value
    tracker    = meta_sheet['B2'].value
    start_date = meta_sheet['B3'].value
    end_date   = meta_sheet['B4'].value

    if project is None:
        print("Project not defined in metadata!")
        exit(1)

    if tracker is None:
        print("Tracker not defined in metadata!")

    if not isinstance(start_date, datetime):
        print("start_date in timesheet metadata of type %s cannot be converted to datetime." % type(start_date))
        exit(1)

    if not isinstance(end_date, datetime):
        print("end_date in timesheet metadata of type %s cannot be converted to datetime." % type(end_date))
        exit(1)

    return TimesheetMetadata(project, tracker, start_date, end_date)

def parse_timesheet(sheet):
    if sheet.title == "META":
        return None

    item_id = sheet.title

    print("Processing item %s..." % item_id)

    rate = sheet['H2'].value

    if not isinstance(rate, float):
        print("Rate \"%s\" of type %s cannot be converted to float - skipping item." % (rate, type(rate)))
        return None

    opened = sheet['H4'].value

    if not isinstance(opened, datetime):
        print("Open date %s of type %s cannot be parsed as a date - skipping item." % (opened, type(opened)))
        return None

    closed = sheet['H5'].value

    if closed is not None and not isinstance(opened, datetime):
        print("Close date %s of type %s cannot be parsed as a date - skipping item." % (closed, type(closed)))
        return None

    if closed is not None and closed < opened:
        print("Open date is after close date - skipping item.")
        return None

    ranges = []

    errored = False

    for row in sheet.rows:
        if row[1].value == "Start":
            continue
        elif row[1].value == "Totals":
            break
        elif row[0].row > sheet.max_row:
            break

        date = row[0].value
        start_time = row[1].value
        end_time = row[2].value

        if date is None:
            continue

        if not isinstance(date, datetime):
            print("Date string %s in row %d is not valid!" % (date, row[0].row))

            errored = True
            continue

        if not isinstance(start_time, time):
            # openpyxl has a bug where midnight is parsed as a datetime representing midnight on 1899-12-30
            if isinstance(start_time, datetime):
                start_time = start_time.time()
            else:
                print("Start time string %s in row %d is not valid!" % (start_time, row[0].row))

                errored = True
                continue

        if not isinstance(end_time, time):
            if isinstance(end_time, datetime):
                end_time = end_time.time()
            else:
                print("End time string %s in row %d is not valid!" % (end_time, row[0].row))

                errored = True
                continue

        start = datetime.combine(date.date(), start_time)
        end = datetime.combine(date.date(), end_time)

        if date < metadata.start_date:
            print("Item %s has labor before cycle start!" % item_id)
        
        if date > metadata.end_date:
            print("Item %s has labor after cycle end!" % item_id)

        if date < opened:
            print("Item %s has labor before open date!" % item_id)

            errored = True
            continue

        if closed is not None and date > closed:
            print("Item %s has labor after close date!" % item_id)

            errored = True
            continue

        ranges.append(TimeRange(start, end))

    if errored:
        print("Skipping item #%s due to range parsing/validation errors." % item_id)
        return None

    return Timesheet(item_id, rate, ranges, opened, closed)

def parse_timesheets(workbook):
    timesheets = []

    for sheet in wb:
        timesheet = parse_timesheet(sheet)

        if timesheet is None:
            continue

        timesheets.append(timesheet)

        print("Read %d ranges for item %s." % (len(timesheet.ranges), timesheet.item))

    return timesheets

parser = argparse.ArgumentParser(description='Utility for copying chunks chunks from one world to another.')
parser.add_argument('workbook', help='Workbook containing timesheet data.')
args = parser.parse_args()

sheet_file = args.workbook

if not path.isfile(sheet_file):
    print("Workbook file %s does not exist!" % sheet_file)
    exit(1)

print("Loading workbook...")

wb = load_workbook(sheet_file)

print("Parsing metadata...")

metadata = parse_metadata(wb)

print("Found metadata for project %s." % metadata.project)

print("Parsing timesheets...")

timesheets = parse_timesheets(wb)

print("Validating time ranges...")

invalid = False

for ts in timesheets:
    r_comb = list(itertools.combinations(ts.ranges, 2))

    for r_pair in r_comb:
        r1 = r_pair[0]
        r2 = r_pair[1]

        if r1.overlaps(r2):
            print("Found overlap within item %s:" % ts.item)
            print("    %s" % r1)
            print("    %s" % r2)
            invalid = True


ts_comb = list(itertools.combinations(timesheets, 2))

for ts_pair in ts_comb:
    ts1 = ts_pair[0]
    ts2 = ts_pair[1]

    for r1 in ts1.ranges:
        for r2 in ts2.ranges:
            if r1.overlaps(r2):
                print("Found overlap between items %s and %s:" % (ts1.item, ts2.item))
                print("    %s" % r1)
                print("    %s" % r2)
                invalid = True

if invalid:
    print("Terminating due to failed range validation.")
    exit(1)

print("All time ranges validated.")

print("Generating invoice...")

total_owed = 0

now = datetime.now()

lines = []

lines.append("# %s Invoice: %s\n" % (metadata.project, now.strftime(DATE_FORMAT_ISO)))
lines.append('\n')
lines.append("This invoice covers labor between %s and %s.\n" % (metadata.start_date.strftime(DATE_FORMAT_ISO), metadata.end_date.strftime(DATE_FORMAT_ISO)))
lines.append('\n')
lines.append('| Item | Opened | Closed | Cost |\n')
lines.append('|---|---|---|---|\n')

for ts in timesheets:
    owed = ts.get_owed()

    openStr = ts.opened.strftime(DATE_FORMAT_ISO)
    closeStr = ts.closed.strftime(DATE_FORMAT_ISO) if ts.closed is not None else "Ongoing"

    tracker_url = metadata.tracker % ts.item

    lines.append("| [%s](%s) | %s | %s | $%.2f\n" % (ts.item, tracker_url, openStr, closeStr, owed))

    total_owed += owed

lines.append("| **Total** | | | $%.2f\n" % total_owed)

invoice_filename = "%s_invoice_%s.md" % (metadata.project.lower().replace(" ", "_"), now.strftime(DATE_FORMAT_FILE_NAME))

with open(invoice_filename, 'w+') as invoice_file:
    invoice_file.writelines(lines)

print("Wrote invoice to %s." % invoice_filename)
